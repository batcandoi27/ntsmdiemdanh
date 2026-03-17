import openpyxl

# Read Excel file
wb = openpyxl.load_workbook('In So Diem Ca Nhan_T9_2025-2026.xlsx', data_only=True)

with open('excel_analysis.txt', 'w', encoding='utf-8') as f:
    f.write("=== SHEETS ===\n")
    f.write(str(wb.sheetnames) + "\n")
    f.write(f"Total sheets: {len(wb.sheetnames)}\n")

    for sheet_name in wb.sheetnames[:5]:  # First 5 sheets
        ws = wb[sheet_name]
        f.write(f"\n=== Sheet: {sheet_name} (rows: {ws.max_row}, cols: {ws.max_column}) ===\n")
        
        for row_num in range(1, min(25, ws.max_row + 1)):
            row_data = []
            for col_num in range(1, min(12, ws.max_column + 1)):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    row_data.append(str(cell_value)[:40])
            if row_data:
                f.write(f"{row_num}: {row_data}\n")

    f.write("\n=== ALL SHEETS SUMMARY ===\n")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f"- {sheet_name}: {ws.max_row} rows\n")

print("Done! Check excel_analysis.txt")
